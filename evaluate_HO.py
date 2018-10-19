# -*- coding: utf-8 -*-
"""
Created on Thu Feb 01 11:37:13 2018

@author: schnepf

script to import Data exported from the labview script fitting KDs.
KDs are associated with their respective sequences

"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# define function to read in sequences from oligo order sheet (created by Script: dinucleotide-Pymutation_wo-doublicates)
def read_in_sequences(filename, number, sheet='Sheet'):
    wb = load_workbook(filename)
    ws = wb[sheet]
    pos = []
    seq = []
    for i in range((number) / 12 + 1):
        for j in range(12):
            if j < 6:
                pos.append(ws['B' + str(i * 12 + 2 + j)].value)
                seq.append(ws['D' + str(i * 12 + 2 + j)].value)
    pos = [x for x in pos if x != None]
    seq = [x for x in seq if x != None]
    result = pd.DataFrame({'Pos': pos, 'Seq': seq})
    return result


def pos_to_number(pos_in, col_n=12, rown_n=8, start_at_zero=False, shift=0):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter = pos_in[0]
    number = pos_in[1:]
    number = int(number)
    row = letters.index(letter)
    if start_at_zero == True:
        return row * col_n + number - 1 + shift
    else:
        return row * col_n + number + shift


def pos_to_seq(DF, number):
    return (DF.Sequence[np.where(DF.Position == number)[0]])


# function that created permutations of dinucleotide, returning numbers of nucleotides if input (consensus) if needed
def dinculeotide(input, cons_number=True):
    letters = ['A', 'C', 'G', 'T']
    din_l = []
    for i in letters:
        for j in letters:
            din = i + j
            if cons_number == True:
                if din != input:
                    din_l.append(din)
                else:
                    din_l.append([letters.index(i), letters.index(j)])
            else:
                din_l.append(din)
    return din_l


# function that created permutations of single mutations, returning numbers of nucleotides if input (consensus) if needed


def assign_KD(KD_mut, KD_mean, seq, start, stop):
    nuc = []
    counter = 0
    N_seq = 0
    stop = stop + 1
    for i in range(start, stop):
        nuc.append(seq[i])
    prae = seq[:start]
    KD_seq = pd.DataFrame(index=range(96), columns=['sequences', 'KD'])
    for pos in range(len(nuc) - 1):

        din = dinculeotide(nuc[pos] + nuc[pos + 1])
        prae_1 = seq[start:start + pos]
        prae_1 = ''.join(prae_1)
        prae_2 = prae + prae_1
        post_1 = seq[start + pos + 2:]
        for k in din:
            if isinstance(k, str):

                KD_seq.set_value(counter, 'sequences', (prae_2 + k + post_1))
                KD_seq.set_value(counter, 'KD', KD_mut[N_seq])
                N_seq += 1
            else:
                KD_seq.set_value(counter, 'sequences', seq)
                KD_seq.set_value(counter, 'KD', KD_mean)
            counter += 1
    return KD_seq


### function to create all dinucleotide permutations
def di_perm(seq, start, stop):
    nuc = []
    stop = stop + 1
    for i in range(start, stop):
        nuc.append(seq[i])
    # post=(seq[stop:])
    prae = seq[:start]
    sequences = []
    for pos in range(len(nuc) - 1):
        din = dinculeotide(nuc[pos] + nuc[pos + 1], False)
        prae_1 = seq[start:start + pos]
        prae_1 = ''.join(prae_1)
        prae_2 = prae + prae_1
        post_1 = seq[start + pos + 2:]
        for k in din:
            sequences.append(prae_2 + k + post_1)
    return sequences


def main(data_path, data_file, dest_name, filename_order, consensus_seq, write_csv,
         pip_scheme='P:\TF-DNA-Binding\FA\Oligos\Higher_order-pip-scheme2.csv', seperator='\t',
         refs=[10, 11, 32, 33, 44, 65, 66, 82], outliers=[], remove_up_low=False, return_refs=False):
    ##read in oligo order
    origin_plates = read_in_sequences(filename_order, 160)
    origin_plate_1 = origin_plates.loc[:47, :]
    origin_plate_2 = origin_plates.loc[48:, :]

    ###read in pipetting scheme to relocate oligos on plate

    pip_file = pd.read_csv(pip_scheme, sep=';')

    pip_file = pip_file.loc[:, 'source':'dest']
    ###define function to translate positions into numbers
    path_total = data_path + data_file
    data = pd.read_csv(path_total, sep=seperator, header=None)
    try:
        if ',' in data[0][0]:
            data = pd.read_csv(path_total, sep=seperator, header=None, decimal=',')
    except:
        pass
    KDs = data[2]
    # pos=range(88)
    true_refs = [x for x in refs if x not in outliers]
    KD_ref = KDs[true_refs]
    if remove_up_low == True:
        KD_ref_sort = sorted(list(KD_ref))
        KD_mean = np.mean(KD_ref_sort[1:-1])
    else:
        KD_mean = np.mean(KD_ref)
    ##create list of KDs for mutations, excluding consensus
    mut = []
    for i in range(len(KDs)):
        if i not in refs and i not in outliers:
            mut.append(i)
    ### match source oligos and destination positions
    top_oligos = pd.DataFrame(columns=['Position', 'Sequence'], index=np.arange(len(origin_plate_2) + 48))
    for i in range(48):
        top_oligos.Position[i] = int(
            pip_file.dest[np.where(pip_file.source == pos_to_number(origin_plate_1.loc[i, 'Pos']))[0]])
        top_oligos.Sequence[i] = origin_plate_1.loc[i, 'Seq']
    for i in range(len(origin_plate_2)):
        top_oligos.Position[i + 48] = int(
            pip_file.dest[np.where(pip_file.source == pos_to_number(origin_plate_2.loc[i + 48, 'Pos'], shift=6))[0]])
        top_oligos.Sequence[i + 48] = origin_plate_2.loc[i + 48, 'Seq']
    KD_mut = KDs[mut].tolist()
    KD_mut_df = pd.concat([top_oligos, pd.DataFrame(KD_mut)], axis=1)
    KD_mut_df.columns = ['Position', 'Sequence', 'KD']
    result_df = KD_mut_df.append({'Position': 12, 'Sequence': consensus_seq, 'KD': KD_mean}, ignore_index=True)
    if write_csv:
        result_df.to_csv(data_path + dest_name)
    if return_refs == True:
        return (result_df, KD_ref)
    else:
        return result_df


if __name__ == '__main__':
    main()
# TODO: add function descriptions, varible descriptions etc.